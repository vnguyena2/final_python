import pandas as pd
import pandas_datareader as pdr
import csv
from datetime import datetime, timedelta
import os.path
import lxml
import xlwings as xw
# from openpyxl import Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.cell.cell import WriteOnlyCell
# from openpyxl import load_workbook
import os.path


#print("hello=")
print(""=:=)
'''
Look up ticker value from the list.
Pause until a continue input is received.
'''
df = pdr.get_data_yahoo('AMZN', start=datetime(2019, 2, 11), end=datetime(2020, 4, 24))
wb = xw.Book('C:/Users/marma/Documents/investments/Positions/Buy-sell/buy-sell-colours_GOLDv17.xlsm')
ws = wb.sheets['Download']
ws.range('C7').value = df
ws.range('B2').value = start=datetime(2020, 2, 11)
ws.range('B4').value = 'AMZN'
ws.range('A1').value = 3
wb.macro('GetData')

wb.save('C:/Users/marma/Documents/investments/Positions/Buy-sell/buy-sell-colours_GOLDv17.xlsm')
from openpyxl import Workbook
# wb = Workbook(write_only=True)
# #wb = load_workbook('C:/Users/marma/Documents/investments/Positions/pandas_openpyxl_v6.xlsx')
# wb = openpyxl.load_workbook('C:/Users/marma/Documents/investments/Positions/Buy-sell/buy-sell-colours_GOLD.xlsm', keep_vba=True)
#
# print(wb.sheetnames)
# #ws = wb.create_sheet()
# ws = wb.active
# #ws.reset_dimensions()
# #ws.delete_cols(1,7)
#
# cell = WriteOnlyCell(ws)
# cell.style = 'Pandas'
#
# def format_first_row(row, cell):
#     for c in row:
#         cell.value = c
#         yield cell
#
# rows = dataframe_to_rows(df)
#
# first_row = format_first_row(next(rows), cell)
#
# ws.append(first_row)
#
# for row in rows:
#     row = list(row)
#     cell.value = row[0]
#     row[0] = cell
#     ws.append(row)
#
# wb.save('C:/Users/marma/Documents/investments/Positions/Buy-sell/buy-sell-colours_GOLDv3.xlsx')

##### Sheet write attempt ############################
# import pyperclip as clp
#
# #Copy dataframe to clipboard
# df.to_clipboard()
# #paste the clipboard to a valirable
# cells = clp.paste()
# #split text in varialble as rows and columns
# cells = [x.split() for x in cells.split('\n')]

#wb = load_workbook(filename='C:/Users/marma/Documents/investments/Positions/Buy-sell/buy-sell-colours_GOLD.xlsm', read_only=False, keep_vba=True)

#print(wb.sheetnames)

#sheet = wb.get_sheet_by_name('spam')
#ws = wb['Download']

## Title sheet####
#ws.title = 'df data'
#Paste clipboard values to the sheet
# for i, r in zip(range(1,len(cells)), cells):
#     for j, c in zip(range(1,len(r)), r):
#         ws.range(cell(row = i + 6, column = j +2)).value = c
#Save the workbook

import pandas as pd
import pandas_datareader as pdr
import csv
from datetime import datetime, timedelta
import os.path
import lxml
import xlwings as xw
# from openpyxl import Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.cell.cell import WriteOnlyCell
# from openpyxl import load_workbook
import os.path
'''
Look up ticker value from the list.
Pause until a continue input is received.
'''
df = pdr.get_data_yahoo('AMZN', start=datetime(2019, 2, 11), end=datetime(2020, 4, 24))
wb = xw.Book('C:/Users/marma/Documents/investments/Positions/Buy-sell/buy-sell-colours_GOLDv17.xlsm')
ws = wb.sheets['Download']
ws.range('C7').value = df
ws.range('B2').value = start=datetime(2020, 2, 11)
ws.range('B4').value = 'AMZN'
ws.range('A1').value = 3
wb.macro('GetData')

wb.save('C:/Users/marma/Documents/investments/Positions/Buy-sell/buy-sell-colours_GOLDv17.xlsm')
from openpyxl import Workbook
# wb = Workbook(write_only=True)
# #wb = load_workbook('C:/Users/marma/Documents/investments/Positions/pandas_openpyxl_v6.xlsx')
# wb = openpyxl.load_workbook('C:/Users/marma/Documents/investments/Positions/Buy-sell/buy-sell-colours_GOLD.xlsm', keep_vba=True)
#
# print(wb.sheetnames)
# #ws = wb.create_sheet()
# ws = wb.active
# #ws.reset_dimensions()
# #ws.delete_cols(1,7)
#
# cell = WriteOnlyCell(ws)
# cell.style = 'Pandas'
#
# def format_first_row(row, cell):
#     for c in row:
#         cell.value = c
#         yield cell
#
# rows = dataframe_to_rows(df)
#
# first_row = format_first_row(next(rows), cell)
#
# ws.append(first_row)
#
# for row in rows:
#     row = list(row)
#     cell.value = row[0]
#     row[0] = cell
#     ws.append(row)
#
# wb.save('C:/Users/marma/Documents/investments/Positions/Buy-sell/buy-sell-colours_GOLDv3.xlsx')

##### Sheet write attempt ############################
# import pyperclip as clp
#
# #Copy dataframe to clipboard
# df.to_clipboard()
# #paste the clipboard to a valirable
# cells = clp.paste()
# #split text in varialble as rows and columns
# cells = [x.split() for x in cells.split('\n')]

#wb = load_workbook(filename='C:/Users/marma/Documents/investments/Positions/Buy-sell/buy-sell-colours_GOLD.xlsm', read_only=False, keep_vba=True)

#print(wb.sheetnames)

#sheet = wb.get_sheet_by_name('spam')
#ws = wb['Download']

## Title sheet####
#ws.title = 'df data'
#Paste clipboard values to the sheet
# for i, r in zip(range(1,len(cells)), cells):
#     for j, c in zip(range(1,len(r)), r):
#         ws.range(cell(row = i + 6, column = j +2)).value = c
#Save the workbook

>>>>>>> 13d7e87080535e6beac1477671e142109247c8b8
